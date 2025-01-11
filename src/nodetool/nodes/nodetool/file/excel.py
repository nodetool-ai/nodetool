from datetime import datetime
import os
from turtle import pd
from typing import Any, List, Optional, Union
from enum import Enum
from pydantic import Field
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from nodetool.workflows.base_node import BaseNode
from nodetool.workflows.processing_context import ProcessingContext
from nodetool.metadata.types import DataframeRef, ExcelRef, FilePath


class CreateWorkbook(BaseNode):
    """
    Creates a new Excel workbook.
    excel, workbook, create

    Use cases:
    - Initialize new Excel files
    - Start spreadsheet creation workflows
    """

    sheet_name: str = Field(
        default="Sheet1", description="Name for the first worksheet"
    )

    async def process(self, context: ProcessingContext) -> ExcelRef:
        wb = openpyxl.Workbook()
        assert wb.active is not None, "Workbook is empty"
        if wb.active.title != self.sheet_name:
            wb.active.title = self.sheet_name
        return ExcelRef(data=wb)


class DataFrameToExcel(BaseNode):
    """
    Writes a DataFrame to an Excel worksheet.
    excel, dataframe, export

    Use cases:
    - Export data analysis results
    - Create reports from data
    """

    workbook: ExcelRef = Field(
        default=ExcelRef(), description="The Excel workbook to write to"
    )
    dataframe: DataframeRef = Field(
        default=DataframeRef(), description="DataFrame to write"
    )
    sheet_name: str = Field(default="Sheet1", description="Target worksheet name")
    start_cell: str = Field(default="A1", description="Starting cell for data")
    include_header: bool = Field(default=True, description="Include column headers")

    async def process(self, context: ProcessingContext) -> Any:
        df = await context.dataframe_to_pandas(self.dataframe)
        workbook = self.workbook.data
        assert isinstance(
            workbook, openpyxl.Workbook
        ), "Workbook is not an instance of openpyxl.Workbook"

        if self.sheet_name not in workbook.sheetnames:
            workbook.create_sheet(self.sheet_name)

        ws = workbook[self.sheet_name]

        # Write headers
        if self.include_header:
            for col, header in enumerate(df.columns, 1):
                ws.cell(row=1, column=col, value=str(header))

        # Write data
        for row_idx, row in enumerate(df.values, 1 if not self.include_header else 2):
            for col_idx, value in enumerate(row, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        return ExcelRef(data=workbook)


class ExcelToDataFrame(BaseNode):
    """
    Reads an Excel worksheet into a pandas DataFrame.
    excel, dataframe, import

    Use cases:
    - Import Excel data for analysis
    - Process spreadsheet contents
    """

    workbook: ExcelRef = Field(
        default=ExcelRef(), description="The Excel workbook to read from"
    )
    sheet_name: str = Field(default="Sheet1", description="Source worksheet name")
    has_header: bool = Field(default=True, description="First row contains headers")

    async def process(self, context: ProcessingContext) -> DataframeRef:
        workbook = self.workbook.data
        assert isinstance(
            workbook, openpyxl.Workbook
        ), "Workbook is not an instance of openpyxl.Workbook"
        ws = workbook[self.sheet_name]
        data = []
        headers = []

        for row in ws.iter_rows(values_only=True):
            if not headers and self.has_header:
                headers = [str(cell) for cell in row]
                continue
            data.append(row)

        df = pd.DataFrame(data, columns=headers if self.has_header else None)
        return await context.dataframe_from_pandas(df)


class FormatCells(BaseNode):
    """
    Applies formatting to a range of cells.
    excel, format, style

    Use cases:
    - Highlight important data
    - Create professional looking reports
    """

    workbook: ExcelRef = Field(
        default=ExcelRef(), description="The Excel workbook to format"
    )
    sheet_name: str = Field(default="Sheet1", description="Target worksheet name")
    cell_range: str = Field(
        default="A1:B10", description="Cell range to format (e.g. 'A1:B10')"
    )
    bold: bool = Field(default=False, description="Make text bold")
    background_color: Optional[str] = Field(
        default=None,
        description="Background color in hex format (e.g. 'FFFF00' for yellow)",
    )
    text_color: Optional[str] = Field(
        default=None, description="Text color in hex format"
    )

    async def process(self, context: ProcessingContext) -> Any:
        workbook = self.workbook.data
        assert isinstance(
            workbook, openpyxl.Workbook
        ), "Workbook is not an instance of openpyxl.Workbook"
        ws = workbook[self.sheet_name]

        font = Font(bold=self.bold, color=self.text_color)
        fill = (
            PatternFill(
                start_color=self.background_color,
                end_color=self.background_color,
                fill_type="solid",
            )
            if self.background_color
            else None
        )

        for row in ws[self.cell_range]:
            for cell in row:
                cell.font = font
                if fill:
                    cell.fill = fill

        return self.workbook


class SaveWorkbook(BaseNode):
    """
    Saves an Excel workbook to disk.
    excel, save, export

    Use cases:
    - Export final spreadsheet
    - Save work in progress
    """

    workbook: ExcelRef = Field(
        default=ExcelRef(), description="The Excel workbook to save"
    )
    filepath: FilePath = Field(
        default=FilePath(),
        description="""
        Path where to save the file.
        You can use time and date variables to create unique names:
        %Y - Year
        %m - Month
        %d - Day
        %H - Hour
        %M - Minute
        %S - Second
        """,
    )

    async def process(self, context: ProcessingContext):
        workbook = self.workbook.data
        assert isinstance(
            workbook, openpyxl.Workbook
        ), "Workbook is not an instance of openpyxl.Workbook"
        assert self.filepath.path, "Path is not set"
        filename = datetime.now().strftime(self.filepath.path)
        expanded_path = os.path.expanduser(filename)
        workbook.save(expanded_path)


class AutoFitColumns(BaseNode):
    """
    Automatically adjusts column widths to fit content.
    excel, format, columns

    Use cases:
    - Improve spreadsheet readability
    - Professional presentation
    """

    workbook: ExcelRef = Field(
        default=ExcelRef(), description="The Excel workbook to format"
    )
    sheet_name: str = Field(default="Sheet1", description="Target worksheet name")

    async def process(self, context: ProcessingContext) -> Any:
        workbook = self.workbook.data
        assert isinstance(
            workbook, openpyxl.Workbook
        ), "Workbook is not an instance of openpyxl.Workbook"
        ws = workbook[self.sheet_name]

        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass

            adjusted_width = max_length + 2
            ws.column_dimensions[column_letter].width = adjusted_width

        return self.workbook
